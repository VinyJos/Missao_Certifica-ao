from turtle import heading
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class Carro: 
   def __init__(self, nome, cpf, nasc, ):
       self.nome = nome
       self.cpf = cpf
       self.nasc = nasc

n = str(input('Digite o nome: '))
m = str(input('Digite o cpf do veículo: '))
a = int(input('Digite o ano do veículo: '))

carro1 = Carro (n, m, a)

#----------------------------------------------
#Para colocar a classe carro na coluna do arquivo em xlsx
wb = load_workbook('lista_carros2.xlsx')
ws = wb['Cars']

ws.append([carro1.nome, carro1.cpf, carro1.nasc])
wb.save('lista_carros2.xlsx')
print(carro1.nome)